import os
import re
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Alignment, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from datetime import datetime
import tempfile

# Creazione di uno stile per il formato delle celle in euro
euro_style = NamedStyle(name="euro_style")
euro_style.number_format = '€ #,##0.00'  # Formato valuta in euro

def italian_month_to_number(month):
    # Dizionario per la conversione dei nomi dei mesi italiani in numeri
    months = {
        "Gennaio": 1,
        "Febbraio": 2,
        "Marzo": 3,
        "Aprile": 4,
        "Maggio": 5,
        "Giugno": 6,
        "Luglio": 7,
        "Agosto": 8,
        "Settembre": 9,
        "Ottobre": 10,
        "Novembre": 11,
        "Dicembre": 12
    }
    return months.get(month, 0)  # Restituisce 0 se il mese non è presente nel dizionario

def nome_mese(mese):
    mesi = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
    return mesi[mese - 1] if 1 <= mese <= 12 else None

def extract_specific_lines_to_excel(pdf_folder):
    # Elenco delle cartelle dei dipendenti nella cartella specificata
    employees = [f.path for f in os.scandir(pdf_folder) if f.is_dir()]
    
    # Itera attraverso ogni cartella (dipendente)
    for employee_folder in employees:
        # Estrai il nome del dipendente dalla cartella
        employee_name = os.path.basename(employee_folder)
        
        # Crea un nuovo foglio di lavoro Excel per il dipendente corrente
        wb = Workbook()
        
        # Dizionario per memorizzare le somme delle righe per ogni legenda
        summary = {}
        
        # Elenco delle cartelle (anni) per il dipendente corrente
        years = [f.path for f in os.scandir(employee_folder) if f.is_dir()]
        
        # Itera attraverso ogni cartella (anno) per il dipendente corrente
        for year_folder in years:
            # Estrai l'anno dalla cartella
            year = os.path.basename(year_folder)
            
            # Crea un nuovo foglio di lavoro Excel per l'anno corrente
            ws = wb.create_sheet(title=year)
            
            # Elenco dei file PDF nelle cartelle del dipendente e dell'anno corrente
            pdf_files = [f for f in os.listdir(year_folder) if f.endswith('.pdf') or f.endswith('.PDF')]
            
            # Dizionario per memorizzare i valori estratti per ogni codice e mese
            data = {}
            
            # Itera attraverso ogni file PDF per il dipendente e l'anno corrente
            for pdf_file in pdf_files:
                # Crea un file di testo temporaneo
                temp_text_file = tempfile.NamedTemporaryFile(mode='w', delete=False)
                
                # Estrai il mese e l'anno dal nome del file PDF
                filename, file_extension = os.path.splitext(pdf_file)
                month, year = filename.split(' ', 1)
                
                # Path completo del file PDF
                pdf_path = os.path.join(year_folder, pdf_file)
                
                # Apre il file PDF con pdfplumber
                with pdfplumber.open(pdf_path) as pdf:
                    # Itera attraverso tutte le pagine del PDF
                    for page_num in range(len(pdf.pages)):
                        page = pdf.pages[page_num]
                        text = page.extract_text()
                        
                        # Itera attraverso tutte le righe del testo
                        for line in text.split('\n'):
                            # Utilizza espressioni regolari per trovare le righe con i codici desiderati
                            match = re.match(r'^(0969|0970|0991|0992|0AD0|0AD1|0421|0457|0131|0576|0376|0377|0169|0170|0965|0966|0967|0987|0988|0790|0076)\s', line)
                            if match:
                                # Scrivi la riga nel file di testo temporaneo
                                temp_text_file.write(line + '\n')
                
                # Chiudi il file di testo temporaneo
                temp_text_file.close()
                
                # Analizza il file di testo temporaneo e aggiorna i dati
                with open(temp_text_file.name, 'r') as temp_file:
                    for line in temp_file:
                        # Separa la riga in codice, descrizione e valore
                        cells = line.split()
                        if len(cells) >= 3:
                            # Sostituisci la virgola con un punto per consentire la conversione in float
                            value = cells[-1].replace(',', '.')
                            if month not in data:
                                data[month] = {}
                            if cells[0] not in data[month]:
                                data[month][cells[0]] = float(value)
                            else:
                                data[month][cells[0]] += float(value)
                
                # Elimina il file di testo temporaneo
                os.remove(temp_text_file.name)
            
            # Ordina i mesi in ordine cronologico utilizzando il numero del mese
            months_sorted = sorted(data.keys(), key=lambda x: italian_month_to_number(x))
            
            # Scrittura dei dati nel foglio di lavoro
            ws.append(['Codice'] + months_sorted)  # Intestazioni delle colonne con i mesi
            for code in sorted(data[months_sorted[0]].keys()):
                row = [code]
                for month in months_sorted:
                    row.append(data[month].get(code, ''))
                ws.append(row)
            
            # Modifica il nome del foglio di lavoro con il nome del dipendente e l'anno lavorativo
            ws.title = f'Anno {year}'
            
            # Somma di tutte le celle da B alla colonna corrente e scrittura nella cella corrispondente alla riga Totale (riga 30)
            last_row = 30  # Righe del totale
            ws[f'O{last_row}'] = f"=SUM(B{last_row}:M{last_row})"
            
            # Scrittura di "TOTALE" nella cella A corrispondente alla riga Totale (riga 30)
            ws[f'A{last_row}'] = "TOTALE"
            
            # Applica lo stile in grassetto alla riga del totale
            ws[f'A{last_row}'].font = Font(bold=True)
            for col in range(2, 14):  # Modifica il range fino alla colonna M
                ws[get_column_letter(col) + str(last_row)].font = Font(bold=True)
                
                # Somma delle righe dalla riga 2 alla riga 29 e scrittura nella riga 30 per ogni colonna
                for col in range(2, 14):  # Modifica il range fino alla colonna M
                    ws[get_column_letter(col) + '30'] = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}29)"
            
            # Aggiungi le scritte richieste nelle celle specifiche
            ws['A31'] = "Presenze"
            ws['A32'] = "Ferie"
            ws['A33'] = "Retribuzione mensile"
            ws['A34'] = "VALORE MEDIO GIORNALIERO VOCI CONTRATTUALI ACCESSORIE"
            ws['A35'] = "RETRIBUZIONE GIORNALIERA (1/26 ART.68, COMM.6 DEL CCNL)"
            ws['A36'] = "INCIDENZA"
            
            # Applica lo stile in grassetto alla riga 34
            ws['A34'].font = Font(bold=True)
            ws['A35'].font = Font(bold=True)
            ws['A36'].font = Font(bold=True)
            # Applicazione dello stile al testo in rosso delle celle specificate
            for cell in ['A36', 'A37', 'B36', 'C36', 'H37']:
                ws[cell].font = Font(color="FF0000")

            # Applicazione dello sfondo giallo alle celle specificate
            ws['A31'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['B31'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['A32'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['B32'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['B33'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['A33'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['C33'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        
            
            
     
            # Modifica il nome del foglio di lavoro con il nome del dipendente e l'anno lavorativo
            ws.title = f'Anno {year}'

            # Calcolo dei valori nelle celle specificate
            ws['G34'] = f"=O30/B31"
            ws['G35'] = f"=C33/26"
            ws['B36'] = f"=(G34*100)/G35"
            ws['C36'] = f"%"
            ws['A37'] = "Voci contrattuali accessorie dovute durante le ferie (valore medio giornaliero x gg di ferie anno)"
            ws['H37'] = f"=B32*G34"
            # Scrivi l'anno delle buste paga nella cella N1
            ws['N1'] = int(year)
            # Applicazione dello stile alle celle specificate
            for cell in ['G34', 'G35']:
                ws[cell].style = euro_style

            # Applicazione dello stile alle celle della riga 30 tranne A30
            for cell in ws[30]:
                if cell.column_letter != 'A':
                    cell.style = euro_style

            for row in range(2, 30):
                for col in range(2, 14):
                    cell = ws.cell(row=row, column=col)
                    cell.style = euro_style

            # Allinea il testo delle celle
            for row in ws.iter_rows(min_row=31, max_row=36, min_col=1, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left')

            # Applicazione dello stile bordi visibili alle celle specificate
            # Definisci lo stile del bordo
            border_style = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))

            # Bordi visibili
            for row in range(1, 30):
                for col in range(1, 14):
                    cell = ws.cell(row=row, column=col)
                    if col <= 13 and cell.value:
                        cell.border = border_style
                        cell.font = Font(bold=False)

            # Bordi per le celle di riga 30 con testo
            for col in range(1, 16):
                cell = ws.cell(row=30, column=col)
                if cell.value:
                    cell.border = border_style
                    cell.font = Font(bold=True)
                
            cell_list = ['A31', 'A32', 'A33', 'B31', 'B32', 'B33', 'C33', 'A34', 'B34', 'C34', 'D34', 'E34', 'F34', 'G34', 'A35', 'B35', 'C35', 'D35', 'E35', 'F35', 'G35', 'A36', 'B36', 'A37', 'B37', 'C37', 'D37', 'E37', 'F37', 'G37', 'H37']
            for cell in cell_list:
                ws[cell].border = border_style

            ws.column_dimensions['O'].width = 10
            
            # Path del file Excel per il dipendente corrente
            excel_path = os.path.join(pdf_folder, f'{employee_name}.xlsx')

            
            # Salva il foglio di lavoro Excel per il dipendente corrente
            wb.save(excel_path)
            
            print(f"Estrazione completata per il dipendente {employee_name} anno {year}.")

# Funzione per preparare i file PDF
def prepare_and_extract(pdf_folder):
    # Attraversa ricorsivamente tutte le sottocartelle e trova i file PDF
    for dirpath, _, filenames in os.walk(pdf_folder):
        for filename in filenames:
            if filename.endswith('.PDF') or filename.endswith('.pdf'):
                file_path = os.path.join(dirpath, filename)
                try:
                    # Estrapola l'anno e il mese dal nome del file
                    anno = int(filename[:4])
                    mese = int(filename[5:7])
                except ValueError:
                    print(f"Il file '{filename}' non segue il formato YYYY_MM.pdf. Ignorato.")
                    continue

                # Ottieni il nome del mese in italiano
                nome_mese_it = nome_mese(mese)

                if nome_mese_it:
                    # Rimuovi spazi extra dal nome del mese
                    nome_mese_it = nome_mese_it.strip()

                    # Rinomina il file con il formato desiderato
                    nuovo_nome = f"{nome_mese_it} {anno}.pdf"
                    # Rinomina il file
                    nuovo_path = os.path.join(dirpath, nuovo_nome)
                    os.rename(file_path, nuovo_path)
                    print(f"File rinominato: {file_path} -> {nuovo_path}")
                else:
                    print(f"Errore: Mese non valido nel file {filename}")
    
    # Dopo la preparazione dei PDF, esegui l'estrazione dei dati
    extract_specific_lines_to_excel(pdf_folder)


# Chiedi all'utente di inserire il percorso della cartella dei dipendenti con le buste paga
customers_folder = input("Inserisci il percorso della cartella dei dipendenti con le buste paga: ")

# Chiama la funzione per preparare i file PDF e poi estrarre le righe specifiche dai PDF e scriverle in fogli Excel per ogni dipendente e anno
prepare_and_extract(customers_folder)

# Itera attraverso tutti i file Excel creati
for root, dirs, files in os.walk(customers_folder):
    for file in files:
        if file.endswith('.xlsx') or file.endswith('.xls'):
            # Apri il file Excel esistente
            excel_path = os.path.join(root, file)
            wb = load_workbook(excel_path)
            # Itera attraverso tutti i file Excel creati
            for root, dirs, files in os.walk(customers_folder):
                for file in files:
                    if file.endswith('.xlsx') or file.endswith('.xls'):
                        # Apri il file Excel esistente
                        excel_path = os.path.join(root, file)
                        wb = load_workbook(excel_path)
                        
                        # Rimuovi il foglio "sheet" se presente
                        if "Sheet" in wb.sheetnames:
                            sheet = wb["Sheet"]
                            wb.remove(sheet)
                        
                        # Crea un nuovo foglio di lavoro "Riepilogo" solo se non esiste già
                        if "Riepilogo" not in wb.sheetnames:
                            ws = wb.create_sheet(title="Riepilogo")
                            # Salva il file Excel con il foglio "Riepilogo"
                            wb.save(excel_path)
                            # Sposta il foglio "Riepilogo" indietro di un numero di posizioni pari al numero di fogli precedenti
                            wb.move_sheet(ws, offset=-len(wb.sheetnames) + 1)
                            
                            # Aggiungi il contenuto desiderato al foglio "Riepilogo"
                            ws['A1'] = "Riepilogo"


                            # Aggiungi i nomi degli altri fogli presenti nel documento excel, in ordine alfabetico
                            sheet_names = sorted(sheet for sheet in wb.sheetnames if sheet != "Riepilogo")
                            for i, sheet_name in enumerate(sheet_names):
                                cell = get_column_letter(i+2) + '2'
                                ws[cell] = sheet_name

                            # Aggiungi la formula nella riga 3 per ogni cella della riga 2 che contiene del testo
                            for col in range(2, ws.max_column + 1):
                                cell = get_column_letter(col) + '3'
                                if ws.cell(row=2, column=col).value:
                                    ws[cell] = f"='{ws.cell(row=2, column=col).value}'!$B$36"
                            
                            # Applicazione dello stile in grassetto e rosso alla cella A3
                            ws['A3'].font = Font(bold=True, color="FF0000")
                            ws['A3'].value = "INCIDENZA"
                            
                            # Applicazione dello stile in grassetto e rosso alla cella A4
                            ws['A4'].font = Font(bold=True, color="FF0000")
                            ws['A4'].value = "Voci contrattuali accesìsorie dovute "
        
                            # Aggiungi la formula nella riga 4 per ogni cella della riga 2 che contiene del testo
                            for col in range(2, ws.max_column + 1):
                                cell = get_column_letter(col) + '4'
                                if ws.cell(row=2, column=col).value:
                                    ws[cell] = f"='{ws.cell(row=2, column=col).value}'!$H$37"
                    
                            # Modifica la larghezza della colonna A a 30 (185px)
                            ws.column_dimensions['A'].width = 30

                            # Aggiungi "TOTALE" in grassetto alla fine della riga 2
                            last_column = get_column_letter(ws.max_column + 1)
                            ws[last_column + '2'].font = Font(bold=True)
                            ws[last_column + '2'].value = "TOTALE"

                            # Calcola la somma delle celle dalla B3 fino all'ultima cella occupata della riga 3
                            last_column = get_column_letter(ws.max_column)
                            ws[last_column + '3'].value = f"=SUM(B3:{last_column}3)"


                            # Modifica la larghezza della colonna corrispondente alla cella con il simbolo "%"
                            ws.column_dimensions[last_column].width = 5
                            
                            # Calcola la somma delle celle dalla B4 fino all'ultima cella occupata della riga 4
                            last_column = get_column_letter(ws.max_column)
                            ws[last_column + '4'].value = f"=SUM(B4:{last_column}4)"

                            # Formatta le celle di riga 4 come valuta in euro
                            for col in range(2, ws.max_column + 1):
                                cell = get_column_letter(col) + '4'
                                ws[cell].style = euro_style
    
                            wb.save(excel_path)

                            # Applicazione dello stile bordi visibili alle celle specificate
                            # Definisci lo stile del bordo
                            border_style = Border(left=Side(style='thin'),
                                                  right=Side(style='thin'),
                                                  top=Side(style='thin'),
                                                  bottom=Side(style='thin'))

                            # Bordi visibili per le celle con del testo
                            for row in range(1, ws.max_row + 1):
                                for col in range(1, ws.max_column + 1):
                                    cell = ws.cell(row=row, column=col)
                                    if cell.value:
                                        cell.border = border_style

                            # Aggiungi il simbolo "%" nella cella dopo l'ultima con del testo nella riga 3
                            last_column = get_column_letter(ws.max_column + 1)
                            ws[last_column + '3'].value = "%"

                            # Modifica la larghezza delle colonne con del testo (eccetto la colonna A) a 10
                            for col in range(2, ws.max_column + 1):
                                column_letter = get_column_letter(col)
                                ws.column_dimensions[column_letter].width = 10

                            
                            # Itera attraverso tutte le celle della riga 4  
                            for col in range(2, ws.max_column + 1):
                                cell = ws.cell(row=4, column=col)
                                cell.font = Font(color="FF0000")  # Imposta il colore del testo a rosso


                            # Ordina i fogli Excel in ordine alfabetico, tranne "Riepilogo" che viene messo per primo
                            wb._sheets.sort(key=lambda x: x.title.lower() if x.title != "Riepilogo" else "")

                            # Salva il file Excel con il foglio "Riepilogo"
                            wb.save(excel_path)

# Aggiungi un premi invio per chiudere il programma
input("Premi invio per chiudere il programma")
